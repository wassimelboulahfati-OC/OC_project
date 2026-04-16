import json
import csv
import os
import datetime
import sys

# --- CONFIGURATION ---
output_dir = "csv_output"
log_dir = "logs"
os.makedirs(output_dir, exist_ok=True)
os.makedirs(log_dir, exist_ok=True)

TEC_DATE = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# --- FONCTIONS UTILITAIRES ---

def write_log(log_path, log_entries):
    """Charge le log existant, ajoute les nouvelles entrées, et réécrit le fichier."""
    log_columns = ["id_station", "dh_utc", "date_traitement", "nature", "detail"]
    existing_logs = []
    if os.path.exists(log_path):
        with open(log_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            existing_logs = list(reader)
    existing_logs.extend(log_entries)
    with open(log_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=log_columns)
        writer.writeheader()
        writer.writerows(existing_logs)
    return len(existing_logs)


def compare_and_log(old_records, new_records, all_columns, tec_date):
    """Compare anciennes et nouvelles données, retourne les entrées de log."""
    log_entries = []
    counters = {"ADDED": 0, "MODIFIED": 0, "REMOVED": 0, "ERROR": 0}

    for key, new_record in new_records.items():
        if key not in old_records:
            log_entries.append({
                "id_station": key[0],
                "dh_utc": key[1],
                "date_traitement": tec_date,
                "nature": "ADDED",
                "detail": ""
            })
            counters["ADDED"] += 1
        else:
            old_record = old_records[key]
            modified = False
            for col in all_columns:
                old_val = old_record.get(col, "")
                new_val = str(new_record.get(col, "")) if new_record.get(col) is not None else ""
                if old_val != new_val:
                    modified = True
                    break
            if modified:
                log_entries.append({
                    "id_station": key[0],
                    "dh_utc": key[1],
                    "date_traitement": tec_date,
                    "nature": "MODIFIED",
                    "detail": ""
                })
                counters["MODIFIED"] += 1

    for key in old_records:
        if key not in new_records:
            log_entries.append({
                "id_station": key[0],
                "dh_utc": key[1],
                "date_traitement": tec_date,
                "nature": "REMOVED",
                "detail": ""
            })
            counters["REMOVED"] += 1

    return log_entries, counters


def load_existing_csv(csv_path):
    """Charge un CSV existant dans un dictionnaire indexé par (id_station, dh_utc)."""
    old_records = {}
    if os.path.exists(csv_path):
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = (row['id_station'], row['dh_utc'])
                old_records[key] = row
        print(f"  Données existantes chargées: {len(old_records)} mesures")
    else:
        print("  Aucune donnée existante, premier chargement")
    return old_records


# --- TRAITEMENT JSON (InfoClimat) ---

def process_json(file_path):
    print(f"\n{'='*60}")
    print(f"TRAITEMENT JSON : {file_path}")
    print(f"{'='*60}")

    csv_path = os.path.join(output_dir, "infoclimat.csv")
    log_path = os.path.join(log_dir, "log_infoclimat.csv")
    log_entries = []

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except FileNotFoundError:
        print(f"  ERREUR: Fichier introuvable - {file_path}")
        log_entries.append({
            "id_station": "N/A",
            "dh_utc": "N/A",
            "date_traitement": TEC_DATE,
            "nature": "ERROR",
            "detail": f"Fichier introuvable: {file_path}"
        })
        write_log(log_path, log_entries)
        return
    except json.JSONDecodeError as e:
        print(f"  ERREUR: JSON invalide - {e}")
        log_entries.append({
            "id_station": "N/A",
            "dh_utc": "N/A",
            "date_traitement": TEC_DATE,
            "nature": "ERROR",
            "detail": f"JSON invalide: {e}"
        })
        write_log(log_path, log_entries)
        return

    # Vérifier la structure du JSON
    if 'hourly' not in raw or 'stations' not in raw:
        print("  ERREUR: Structure JSON inattendue (clés 'hourly' ou 'stations' manquantes)")
        log_entries.append({
            "id_station": "N/A",
            "dh_utc": "N/A",
            "date_traitement": TEC_DATE,
            "nature": "ERROR",
            "detail": "Structure JSON inattendue"
        })
        write_log(log_path, log_entries)
        return

    # Affichage des stations
    print("  Stations disponibles:", [s['id'] + ' - ' + s['name'] for s in raw['stations']])

    # Collecte colonnes et mesures
    all_columns = set()
    all_records = []

    for station_id, records in raw['hourly'].items():
        if station_id != '_params':
            if len(records) == 0:
                log_entries.append({
                    "id_station": station_id,
                    "dh_utc": "N/A",
                    "date_traitement": TEC_DATE,
                    "nature": "ERROR",
                    "detail": f"Station {station_id} sans mesures"
                })
                print(f"  ATTENTION: Station {station_id} sans mesures")
                continue
            print(f"  {station_id}: {len(records)} mesures")
            all_columns.update(records[0].keys())
            all_records.extend(records)

    if len(all_records) == 0:
        print("  ERREUR: Aucune mesure trouvée")
        log_entries.append({
            "id_station": "N/A",
            "dh_utc": "N/A",
            "date_traitement": TEC_DATE,
            "nature": "ERROR",
            "detail": "Aucune mesure trouvée dans le fichier"
        })
        write_log(log_path, log_entries)
        return

    print(f"  Total mesures: {len(all_records)}")

    # Charger les anciennes données
    old_records = load_existing_csv(csv_path)

    # Construire le dictionnaire des nouvelles données
    new_records = {}
    for record in all_records:
        try:
            key = (record['id_station'], record['dh_utc'])
            new_records[key] = record
        except KeyError as e:
            log_entries.append({
                "id_station": record.get('id_station', 'N/A'),
                "dh_utc": record.get('dh_utc', 'N/A'),
                "date_traitement": TEC_DATE,
                "nature": "ERROR",
                "detail": f"Clé manquante: {e}"
            })

    # Comparer et générer les logs
    change_logs, counters = compare_and_log(old_records, new_records, all_columns, TEC_DATE)
    log_entries.extend(change_logs)

    # Export CSV
    sorted_columns = sorted(all_columns)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=sorted_columns, extrasaction='ignore')
        writer.writeheader()
        for record in all_records:
            clean_record = {k: (str(v) if v is not None else "") for k, v in record.items()}
            writer.writerow(clean_record)

    print(f"  CSV exporté: {len(all_records)} lignes, {len(sorted_columns)} colonnes")

    # Export log
    total_logs = write_log(log_path, log_entries)

    print(f"\n  --- Résumé ---")
    print(f"  ADDED:    {counters['ADDED']}")
    print(f"  MODIFIED: {counters['MODIFIED']}")
    print(f"  REMOVED:  {counters['REMOVED']}")
    print(f"  ERRORS:   {len([l for l in log_entries if l['nature'] == 'ERROR'])}")
    print(f"  Log total: {total_logs} entrées")


# --- TRAITEMENT XLSX (Weather Underground) ---

def process_xlsx(file_path):
    print(f"\n{'='*60}")
    print(f"TRAITEMENT XLSX : {file_path}")
    print(f"{'='*60}")

    # Déterminer le nom de la station à partir du nom de fichier
    filename = os.path.basename(file_path)
    if "La Madeleine" in filename or "La+Madeleine" in filename:
        station_name = "La Madeleine, FR"
        station_id = "WU_LA_MADELEINE"
    elif "Ichtegem" in filename:
        station_name = "Ichtegem, BE"
        station_id = "WU_ICHTEGEM"
    else:
        station_name = filename.replace(".xlsx", "")
        station_id = "WU_" + station_name.upper().replace(" ", "_").replace(",", "")

    csv_filename = f"weather_underground_{station_id.lower()}.csv"
    csv_path = os.path.join(output_dir, csv_filename)
    log_path = os.path.join(log_dir, f"log_{station_id.lower()}.csv")
    log_entries = []

    try:
        import openpyxl
    except ImportError:
        print("  ERREUR: openpyxl non installé. Exécutez: pip install openpyxl")
        log_entries.append({
            "id_station": station_id,
            "dh_utc": "N/A",
            "date_traitement": TEC_DATE,
            "nature": "ERROR",
            "detail": "Module openpyxl non installé"
        })
        write_log(log_path, log_entries)
        return

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        print(f"  ERREUR: Fichier introuvable - {file_path}")
        log_entries.append({
            "id_station": station_id,
            "dh_utc": "N/A",
            "date_traitement": TEC_DATE,
            "nature": "ERROR",
            "detail": f"Fichier introuvable: {file_path}"
        })
        write_log(log_path, log_entries)
        return
    except Exception as e:
        print(f"  ERREUR: Impossible de lire le fichier - {e}")
        log_entries.append({
            "id_station": station_id,
            "dh_utc": "N/A",
            "date_traitement": TEC_DATE,
            "nature": "ERROR",
            "detail": f"Erreur lecture fichier: {e}"
        })
        write_log(log_path, log_entries)
        return

    print(f"  Station: {station_name} ({station_id})")
    print(f"  Onglets: {wb.sheetnames}")

    all_records = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 3:
            log_entries.append({
                "id_station": station_id,
                "dh_utc": "N/A",
                "date_traitement": TEC_DATE,
                "nature": "ERROR",
                "detail": f"Onglet {sheet_name} vide ou insuffisant"
            })
            print(f"  ATTENTION: Onglet {sheet_name} ignoré (pas assez de lignes)")
            continue

        # La première ligne contient les en-têtes
        headers_raw = rows[0]
        headers = []
        for h in headers_raw:
            if h is not None:
                headers.append(str(h).strip())
            else:
                headers.append("")

        # Convertir le nom de l'onglet en date (ex: 011024 -> 2024-10-01)
        try:
            day = sheet_name[:2]
            month = sheet_name[2:4]
            year = "20" + sheet_name[4:6]
            date_str = f"{year}-{month}-{day}"
            # Vérifier que la date est valide
            datetime.datetime.strptime(date_str, "%Y-%m-%d")
        except (ValueError, IndexError):
            log_entries.append({
                "id_station": station_id,
                "dh_utc": "N/A",
                "date_traitement": TEC_DATE,
                "nature": "ERROR",
                "detail": f"Nom d'onglet invalide pour date: {sheet_name}"
            })
            print(f"  ATTENTION: Onglet {sheet_name} - format de date invalide")
            continue

        # Parcourir les lignes de données (à partir de la ligne 3, index 2)
        row_count = 0
        for row in rows[2:]:
            values = list(row)

            # Ignorer les lignes vides
            if all(v is None for v in values):
                continue

            time_val = values[0] if len(values) > 0 else None
            if time_val is None:
                continue

            # Construire le timestamp complet
            time_str = str(time_val).strip()
            dh_utc = f"{date_str} {time_str}"

            # Créer l'enregistrement
            record = {
                "id_station": station_id,
                "dh_utc": dh_utc,
                "date": date_str,
            }

            # Mapper chaque colonne
            for i, header in enumerate(headers):
                if i < len(values) and header != "":
                    val = values[i]
                    record[header] = str(val).strip() if val is not None else ""

            all_records.append(record)
            row_count += 1

        print(f"  Onglet {sheet_name} ({date_str}): {row_count} mesures")

    if len(all_records) == 0:
        print("  ERREUR: Aucune mesure extraite")
        log_entries.append({
            "id_station": station_id,
            "dh_utc": "N/A",
            "date_traitement": TEC_DATE,
            "nature": "ERROR",
            "detail": "Aucune mesure extraite du fichier"
        })
        write_log(log_path, log_entries)
        return

    print(f"  Total mesures: {len(all_records)}")

    # Collecter toutes les colonnes
    all_columns = set()
    for record in all_records:
        all_columns.update(record.keys())

    print(f"  Colonnes: {sorted(all_columns)}")

    # Charger les anciennes données
    old_records = load_existing_csv(csv_path)

    # Construire le dictionnaire des nouvelles données
    new_records = {}
    for record in all_records:
        try:
            key = (record['id_station'], record['dh_utc'])
            new_records[key] = record
        except KeyError as e:
            log_entries.append({
                "id_station": record.get('id_station', 'N/A'),
                "dh_utc": record.get('dh_utc', 'N/A'),
                "date_traitement": TEC_DATE,
                "nature": "ERROR",
                "detail": f"Clé manquante: {e}"
            })

    # Comparer et générer les logs
    change_logs, counters = compare_and_log(old_records, new_records, all_columns, TEC_DATE)
    log_entries.extend(change_logs)

    # Export CSV
    sorted_columns = sorted(all_columns)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=sorted_columns, extrasaction='ignore')
        writer.writeheader()
        for record in all_records:
            clean_record = {k: (str(v) if v is not None else "") for k, v in record.items()}
            writer.writerow(clean_record)

    print(f"  CSV exporté: {len(all_records)} lignes, {len(sorted_columns)} colonnes")

    # Export log
    total_logs = write_log(log_path, log_entries)

    print(f"\n  --- Résumé ---")
    print(f"  ADDED:    {counters['ADDED']}")
    print(f"  MODIFIED: {counters['MODIFIED']}")
    print(f"  REMOVED:  {counters['REMOVED']}")
    print(f"  ERRORS:   {len([l for l in log_entries if l['nature'] == 'ERROR'])}")
    print(f"  Log total: {total_logs} entrées")

    wb.close()


# --- MAIN ---

if __name__ == "__main__":
    print(f"Début du traitement - {TEC_DATE}")
    print(f"Dossier de travail: {os.getcwd()}")

    if len(sys.argv) < 2:
        print("\nUsage: python prepare_data.py <fichier1> <fichier2> ...")
        print("  Fichiers supportés: .json, .xlsx")
        print("\nExemple:")
        print("  python prepare_data.py Data_Source1.json WU_LaMadeleine.xlsx WU_Ichtegem.xlsx")
        sys.exit(1)

    for file_path in sys.argv[1:]:
        if not os.path.exists(file_path):
            print(f"\nERREUR: Fichier introuvable - {file_path}")
            continue

        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".json":
            process_json(file_path)
        elif ext == ".xlsx":
            process_xlsx(file_path)
        else:
            print(f"\nERREUR: Format non supporté - {ext} ({file_path})")

    print(f"\n{'='*60}")
    print(f"Traitement terminé - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"CSV générés dans: {os.path.abspath(output_dir)}")
    print(f"Logs dans: {os.path.abspath(log_dir)}")

    # --- COPIE VERS AIRBYTE (node kind) ---
    print(f"\n{'='*60}")
    print("Copie des CSV vers Airbyte (node kind)...")
    import subprocess
    csv_files = [
        "infoclimat.csv",
        "weather_underground_wu_la_madeleine.csv",
        "weather_underground_wu_ichtegem.csv",
    ]
    airbyte_node = "airbyte-abctl-control-plane"
    airbyte_dest = "/var/local-path-provisioner/airbyte-local-pv/"
    all_ok = True
    for csv_file in csv_files:
        src = os.path.join(os.path.abspath(output_dir), csv_file)
        if not os.path.exists(src):
            print(f"  IGNORÉ (fichier absent): {csv_file}")
            continue
        result = subprocess.run(
            ["docker", "cp", src, f"{airbyte_node}:{airbyte_dest}"],
            capture_output=True, text=True
        )
        if result.returncode == 0:
            print(f"  OK: {csv_file} copié vers {airbyte_node}:{airbyte_dest}")
        else:
            print(f"  ERREUR copie {csv_file}: {result.stderr.strip()}")
            all_ok = False
    if all_ok:
        print("Copie Airbyte terminée avec succès.")
    else:
        print("Copie Airbyte terminée avec des erreurs (Docker en cours d'exécution ?)")
